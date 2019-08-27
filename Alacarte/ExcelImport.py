from datetime import datetime

from openpyxl import load_workbook

from Alacarte.AlacarteMealContainer import MealContainer
from Alacarte.AlacarteMealDetailsContainer import MealDetailsContainer
from Alacarte.AlacarteMealMenuContainer import MealMenuContainer
from Config import Config
from MysqlDatabase import MysqlDatabase


class ExcelImport:
    def __init__(self, excelPath = Config.alacarteMenuExcelPath):
         self.excelPath = excelPath

    def updateAlacarteMenu(self):
        self.allMealsInFile = []
        self._openExcelFile()
        self._parseOpenedExcelFile()
        self._importAllMealsToDb()


    def _openExcelFile(self):
        self.wb = load_workbook(self.excelPath)
        self.ws = self.wb.active

    def _parseOpenedExcelFile(self):
        #data exists in a -> n

        rowNumber = 2
        parsedMeal = self._iterateRow(rowNumber, "dictionary")
        while parsedMeal != None:
            self.allMealsInFile.append(parsedMeal)
            rowNumber += 1
            parsedMeal = self._iterateRow(rowNumber, "dictionary")

    def _iterateRow(self, rowNumber, iterationMethod):
        valid = self._IsRowValid(rowNumber)
        if not valid:
            return None
        else:
            pass

        if iterationMethod == "dictionary":

            meal = {}
            day = int(self.ws['A' + str(rowNumber)].value)
            month = int(self.ws['B' + str(rowNumber)].value)
            year = int(self.ws['C' + str(rowNumber)].value)
            start_time = self.ws['D' + str(rowNumber)].value
            end_time = self.ws['E' + str(rowNumber)].value

            meal['start_date'] = datetime(year = year, month = month, day = day, hour = start_time.hour, minute = start_time.minute)
            meal['end_date'] = datetime(year = year, month = month, day = day, hour = end_time.hour, minute = end_time.minute)
            meal['tr_type'] = self.ws['F'+str(rowNumber)].value
            meal['en_type'] = self.ws['G'+str(rowNumber)].value
            meal['tr_name'] = self.ws['H'+str(rowNumber)].value
            meal['en_name'] = self.ws['I'+str(rowNumber)].value
            meal['protein'] = self.ws['J'+str(rowNumber)].value
            meal['calorie'] = self.ws['K'+str(rowNumber)].value
            meal['food_type'] = self.ws['L'+str(rowNumber)].value
            
            return meal

        
    def _IsRowValid(self, rowNumber):
        if self.ws['A'+str(rowNumber)].value == None:
            return False
        else:
            return True

    def _importAllMealsToDb(self):
        from MysqlDatabase import MysqlDatabase

        #MongoDatabase().setCafeteriaMenu(self.allMealsInFile)

        MysqlDatabase().setAlacarteMenu(self.allMealsInFile)
