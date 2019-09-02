from datetime import datetime

from openpyxl import load_workbook

from Cafeteria.MealContainer import MealContainer
from Cafeteria.MealDetailsContainer import MealDetailsContainer
from Cafeteria.MealMenuContainer import MealMenuContainer
from Config import Config
from MongoDatabase import MongoDatabase


class AlacarteExcelImport:
    def __init__(self, excelPath = Config.cafeteriaMenuExcelPath):
         self.excelPath = excelPath

    def updateAlacarteMenu(self):
        self.allMealsInFile = []
        self._openExcelFile()
        self._parseOpenedExcelFile()
        self._importAllMealsToDb()


    def _openExcelFile(self):
        self.wb = load_workbook(self.excelPath)
        self.ws = self.wb.active

    def _parseOpenedExcelFile(self):
        #data exists in a -> n

        rowNumber = 2
        parsedMeal = self._iterateRow(rowNumber, "dictionary")
        while parsedMeal != None:
            self.allMealsInFile.append(parsedMeal)
            rowNumber += 1
            parsedMeal = self._iterateRow(rowNumber, "dictionary")

    def _iterateRow(self, rowNumber, iterationMethod):
        valid = self._IsRowValid(rowNumber)
        if not valid:
            return None
        else:
            pass

        if iterationMethod == "dictionary":

            meal = {}
            day = int(self.ws['A' + str(rowNumber)].value)
            month = int(self.ws['B' + str(rowNumber)].value)
            year = int(self.ws['C' + str(rowNumber)].value)
            start_time = self.ws['D' + str(rowNumber)].value
            end_time = self.ws['E' + str(rowNumber)].value

            meal['startTime'] = str(year) + '-' + str(month) + '-' + str(day) + ' ' + str(start_time)
            meal['endTime'] = str(year) + '-' + str(month) + '-' + str(day) + ' ' + str(end_time)
            meal['tr_type'] = self.ws['F'+str(rowNumber)].value
            meal['en_type'] = self.ws['G'+str(rowNumber)].value
            meal['tr_name'] = self.ws['H'+str(rowNumber)].value
            meal['en_name'] = self.ws['I'+str(rowNumber)].value
            meal['calorie'] = self.ws['J'+str(rowNumber)].value
            meal['protein'] = self.ws['K'+str(rowNumber)].value
            meal['food_type'] = self.ws['L'+str(rowNumber)].value
            
            return meal

        elif iterationMethod == "asObject":
            day = int(self.ws['A'+str(rowNumber)].value)
            month = int(self.ws['B'+str(rowNumber)].value)
            year = int(self.ws['C'+str(rowNumber)].value)
            start_time = self.ws['D'+str(rowNumber)].value
            end_time = self.ws['E'+str(rowNumber)].value

            trMealMenuContainer = MealMenuContainer(
                main=self.ws['J'+str(rowNumber)].value,
                pastaOrRice=self.ws['L'+str(rowNumber)].value,
                soup=self.ws['H'+str(rowNumber)].value,
                extra1=self.ws['N'+str(rowNumber)].value,
                extra2=self.ws['P'+str(rowNumber)].value,
                extra3=None
            )

            enMealMenuContainer = MealMenuContainer(
                main=self.ws['K'+str(rowNumber)].value,
                pastaOrRice=self.ws['M'+str(rowNumber)].value,
                soup=self.ws['I'+str(rowNumber)].value,
                extra1=self.ws['O'+str(rowNumber)].value,
                extra2=self.ws['Q'+str(rowNumber)].value,
                extra3=None
            )

            mealDetailsContainer = MealDetailsContainer(
                start=datetime(year = year, month = month, day = day, hour = start_time.hour, minute = start_time.minute),
                end=datetime(year = year, month = month, day = day, hour = end_time.hour, minute = end_time.minute),
                en_name=self.ws['G'+str(rowNumber)].value,
                tr_name=self.ws['F'+str(rowNumber)].value,
                en_menu=enMealMenuContainer,
                tr_menu=trMealMenuContainer
            )
            meal = MealContainer(
                mealDetailsContainer=mealDetailsContainer,
                date=mealDetailsContainer.end
            )

            return meal


    def _IsRowValid(self, rowNumber):
        if self.ws['A'+str(rowNumber)].value == None:
            return False
        else:
            return True

    def _importAllMealsToDb(self):
        from MysqlDatabase import MysqlDatabase

        #MongoDatabase().setCafeteriaMenu(self.allMealsInFile)

        MysqlDatabase().setCafeteriaMenu(self.allMealsInFile)
