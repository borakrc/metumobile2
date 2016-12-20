

class ExcelExport:
    def __init__(self, jsonUrl, nameOfArrayVariableToListInExcelExportFile):
        self.sourceJsonUrl = jsonUrl
        self.nameOfArrayVariableToListInExcelExportFile = nameOfArrayVariableToListInExcelExportFile
        self.arrayToBePutInExcelExportFile = None
        pass
    
    def do(self):
        from datetime import datetime
        self._getDataFromUrl()
        dest_filename = "Excel_Export_" + str(datetime.now().isoformat()) + ".xlsx"
        xlsxExportDestination = self._putDataIntoExcel(data=self.arrayToBePutInExcelExportFile, dest_filename=dest_filename)
        return xlsxExportDestination

    def _getDataFromUrl(self):
        import urllib, json
        url = self.sourceJsonUrl

        response = urllib.urlopen(url).read()
        jsonEndpointData = json.loads(response)

        self.arrayToBePutInExcelExportFile = jsonEndpointData[self.nameOfArrayVariableToListInExcelExportFile]

    def _putDataIntoExcel(self, data, dest_filename):
        dest_filename = dest_filename.replace(":", ".")
        from openpyxl import Workbook
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        dictionaryKeys = self._findDictionaryKeys(data[0])

        column = "A"
        for key in dictionaryKeys:
            ws[column+"1"] = key
            column = str(chr(ord(column) + 1))

        for row in data:
            cellsInARow = []
            for key in dictionaryKeys:
                cellsInARow.append(row[key])

            ws.append(cellsInARow)
            print cellsInARow

        # Save the file
        from Config import Config
        import os
        import openpyxl

        dest_directory = os.path.join(Config.home, 'dynamicFiles')

        self._createDirectoryAndFileIfDoesntExist(directory = dest_directory, filename = dest_filename)

        wb.save(os.path.join(dest_directory, dest_filename))

        return os.path.join(dest_directory, dest_filename)


    def _findDictionaryKeys(self, dictionary):
        assert isinstance(dictionary, dict)

        keyList = []
        for key in dictionary.iterkeys():
            keyList.append(key)

        return keyList


    def _createDirectoryAndFileIfDoesntExist(self, directory, filename):
        import os, openpyxl

        if not os.path.isdir(directory):
            os.makedirs(directory)

        if not os.path.isfile(filename):
            wb = openpyxl.Workbook()
            wb.save(os.path.join(directory, filename))

